from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from tkinter import ttk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd

class GUI():
    def __init__(self,Window):
        self.Window = Window
        self.parent_bg = self.Window.cget('bg')
        self.canvas = tk.Canvas(self.Window)

    def Create_Label(self,Message,X_Pos,Y_Pos,Color = 'black', Font = ('Arial',25),JustifyText='center',AnchorText='nw'):
        tk.Label(self.Window, text=Message, fg= Color , bg=self.parent_bg , font= Font,justify=JustifyText).place(x=X_Pos, y=Y_Pos ,anchor=AnchorText)

    def Create_ComboBox(self,ValuesList,X_Pos,Y_Pos,WidthSize=25):
        List = ttk.Combobox(self.Window, values=ValuesList,width=WidthSize,height=23)
        List.place(x=X_Pos, y=Y_Pos, anchor='nw')
        try:
            List.set(ValuesList[0])
        except:
            pass
        return List

    def Create_Button(self,Message,X_Pos,Y_Pos,Font = ('Arial',25),Function=None):
        Button = tk.Button(self.Window, font= Font, text=Message, width=10, command=Function)
        Button.place(x=X_Pos, y=Y_Pos, anchor='nw')
        return Button
    
    def Create_Rectangle(self, x1, y1, x2, y2, outline='black', width=1):
        self.canvas.create_rectangle(x1, y1, x2, y2, outline=outline, fill=self.parent_bg, width=width)
        self.canvas.pack()

    def EnableWidgets(self,Widgets,Enable):
        if type(Widgets) is list:
            for Widget in Widgets:
                print(Widget)
                Widget["state"] = Enable
        else:
            Widgets["state"] = Enable

    def Select_File(self):
        self.PathFile = filedialog.askopenfilename(title="Select a File", filetype=(('Excel files','*.xlsx'),('all files','*.*')))
        
        last_backslash_index = self.PathFile.rfind("/")
        if last_backslash_index != -1:  # If "\" is found in the string
            self.Directory = self.PathFile[:last_backslash_index] + "/" # Extract the portion of the string up to the last "\"

    def Select_Directory(self):
        self.Directory = filedialog.askdirectory()        
        

class BOM_Template():
    def __init__(self,Path):
        self.Workbook = load_workbook(Path)

    def DeleteTabs(self):
        # Get the list of existing sheet names
        SheetsList = self.Workbook.sheetnames

        # Iterate through the sheets in reverse order to avoid index issues
        for SheetName in reversed(SheetsList):
            Sheet = self.Workbook[SheetName]
            
            # Check if the sheet name is not "AVL Report"
            if SheetName != "AVL Report":
                # Remove the sheet
                self.Workbook.remove(Sheet)
            else:
                worksheet = self.Workbook["AVL Report"]
                worksheet.title = "BOM"

        # Get the list of existing sheet names
        self.BOM_Sheet = self.Workbook['BOM']

    def DeleteColumn(self):
        # Define the columns to delete
        columns_to_delete = ['C', 'D', 'G', 'H', 'J', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']

        # Iterate through the columns in reverse order to avoid index issues
        for column in reversed(columns_to_delete):
            # Get the column index
            column_index = column_index_from_string(column)
            
            # Delete the column
            self.BOM_Sheet.delete_cols(column_index)

    def Change_ColumnWidth(self):

        # Change size of columns
        column_widths = {
            'A': 5.56,
            'B': 13.67,
            'C': 108.90,
            'D': 7.89,
            'E': 41.33,
            'F': 29.33,
            'G': 34.56
        }

        # Set the column widths
        for column, width in column_widths.items():
            self.BOM_Sheet.column_dimensions[column].width = width

    def Change_RowHeight(self):
        # Define the row heights
        row_heights = {
            1: 23.4,
            2: 108,
            3: 23.4,
            4: 23.4,
            5: 23.4,
            6: 23.4,
            7: 23.4,
            8: 23.4,
            9: 23.4
        }

        # Set the row heights
        for row, height in row_heights.items():
            self.BOM_Sheet.row_dimensions[row].height = height

    def AddHeader(self):
        EATON_DESC = self.BOM_Sheet['A3'].value
        self.BOM_Sheet.unmerge_cells("A3:B3")
        self.BOM_Sheet.delete_rows(1,5)
        self.BOM_Sheet.insert_rows(1,9)
        self.BOM_Sheet.merge_cells("A2:D2")
        self.BOM_Sheet.merge_cells("E1:E2")

        self.BOM_Sheet['A2'] = EATON_DESC
        self.BOM_Sheet['C3'] = "FILENAME"
        self.BOM_Sheet['D3'] = "REV"
        self.BOM_Sheet['E3'] = "DESCRIPTION"
        self.BOM_Sheet['E4'] = "BOM"
        self.BOM_Sheet['E5'] = "ASSEMBLY,SCHEMATIC"
        self.BOM_Sheet['E6'] = "PCB FILE"
        self.BOM_Sheet['E7'] = "PCB FAB"
        self.BOM_Sheet['E8'] = "PICK & PLACE FILE"
        self.BOM_Sheet['F1'] = "PICK & PLACE FILE"
        self.BOM_Sheet['E9'] = "HANDLING AND PROCESSING SPECS"
        self.BOM_Sheet['F1'] = "Description:"
        self.BOM_Sheet['E1'] = "Part Name: \nRevision: \nDesigner: \nEngineer: \nDate: \nCO #:"

        # Specify the start and end cells of the range
        start_cell = 'A1'
        end_cell = 'G9'

        # Wrap text in the specified range
        for row in self.BOM_Sheet[start_cell:end_cell]:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

        # Specify the picture name and desired dimensions
        picture_name = 'Picture 1'
        width_inches = 1.25
        height_inches = 0.29

    def merge_cells_by_content(self):
        # Specify the starting row and column
        start_row = 7
        columns = ['A','B']

        # Initialize variables
        current_value = None
        merge_start = start_row

        # Iterate over the cells in the specified column
        for row in range(start_row, self.BOM_Sheet.max_row + 1):
            for column in columns:
                cell_value = self.BOM_Sheet[column + str(row)].value

                if cell_value != current_value:
                    # Check if a merge is needed
                    if merge_start < row - 1:
                        merge_end = row - 1
                        self.BOM_Sheet.merge_cells(f'{column}{merge_start}:{column}{merge_end}')

                    # Update current value and merge start
                    current_value = cell_value
                    merge_start = row

    def merge_empty_cells(self):
        # Specify the starting row and column
        start_row = 7
        columns = ['C','D','E']

        merge_start = None
        merge_end = None

        # Iterate over the cells in the specified column
        for row in range(start_row, self.BOM_Sheet.max_row + 1):
            for column in columns:
                cell_value = self.BOM_Sheet[column + str(row)].value

                if cell_value is None:
                    if merge_start is None:
                        merge_start = row
                    merge_end = row
                elif merge_start is not None:
                    merge_range = f'{column}{merge_start}:{column}{merge_end}'
                    self.BOM_Sheet.merge_cells(merge_range)
                    merge_start = None
                    merge_end = None

    def SaveDocument(self, Path):
        self.Workbook.save(Path)

def TestHMI():
    HMI = tk.Tk()
    HMI.geometry("885x575")
    HMI.title("BOM Template")
    HMI.resizable(0,0) # this removes the maximize button
    HMI.configure(bg='#444444')
    #HMI.iconbitmap("Bingo_Icon.ico")

    Widgets = GUI(HMI)
    Widgets.Create_Button("Open directory",0,0,('Arial',10),Widgets.Select_File)
    HMI.mainloop()

def TestExcel():
    # Load the Excel file
    Path = 'C:/Users/E0619889/Desktop/TestPyexcel/test.xlsx'
    
    Excel = BOM_Template(Path)
    Excel.DeleteTabs()
    Excel.DeleteColumn()
    Excel.Change_ColumnWidth()
    Excel.Change_RowHeight()
    Excel.AddHeader()
    Excel.merge_cells_by_content()
    Excel.merge_empty_cells()

    # Save the modified workbook
    #workbook.save('test_modified.xlsx')
    Excel.SaveDocument('C:\\Users\\E0619889\\Desktop\\TestPyexcel\\test_modified.xlsx')
    print('Done')

if __name__ == '__main__':
    #TestHMI()
    TestExcel()