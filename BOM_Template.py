from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment

class Template():
    def __init__(self,Path):
        self.Workbook = load_workbook(Path)

    def DeleteTabs(self):
        # Get the list of existing sheet names
        SheetsList = self.Workbook.sheetnames

        # Iterate through the sheets in reverse order to avoid index issues
        for SheetName in reversed(SheetsList):
            Sheet = self.Workbook[SheetName]
            
            # Check if the sheet name is not "AVL Report"
            if SheetName != "AVL Report":
                # Remove the sheet
                self.Workbook.remove(Sheet)
            else:
                worksheet = self.Workbook["AVL Report"]
                worksheet.title = "BOM"

        # Get the list of existing sheet names
        self.BOM_Sheet = self.Workbook['BOM']

    def DeleteColumns(self):

        content_list = ['F / N', 'Name', 'Description', 'Quantity', 'Reference\nDesignator', 'MEP Name', 'MEP Manufacturer']

        # Get the max column index
        max_column_index = self.BOM_Sheet.max_column

        # Iterate through the columns and check content in row 7
        columns_to_delete = []
        for i in range(1, max_column_index + 1):
            cell_value = self.BOM_Sheet.cell(row=6, column=i).value
            if cell_value not in content_list:
                columns_to_delete.append(get_column_letter(i))

        # Iterate through the columns to be deleted and remove them
        for column in reversed(columns_to_delete):
            column_index = column_index_from_string(column)
            self.BOM_Sheet.delete_cols(column_index)

    def Change_ColumnWidth(self):
        # Change size of columns
        column_widths = {
            'A': 12.4,
            'B': 18.4,
            'C': 50.4,
            'D': 11.4,
            'E': 42.4,
            'F': 29.4,
            'G': 39.4
        }

        # Set the column widths
        for column, width in column_widths.items():
            self.BOM_Sheet.column_dimensions[column].width = width

    def Change_RowHeight(self):
        # Define the row heights
        row_heights = {
            1: 23.4,
            2: 108,
            3: 23.4,
            4: 23.4,
            5: 23.4,
            6: 23.4,
            7: 23.4,
            8: 23.4,
            9: 23.4,
            10: 30
        }

        # Set the row heights
        for row, height in row_heights.items():
            self.BOM_Sheet.row_dimensions[row].height = height

    def AddHeader_temp(self):
        EATON_DESC = self.BOM_Sheet['A3'].value
        self.BOM_Sheet.unmerge_cells("A3:B3")
        self.BOM_Sheet.delete_rows(1,5)
        self.BOM_Sheet.insert_rows(1,9)
        self.BOM_Sheet.merge_cells("A2:D2")
        self.BOM_Sheet.merge_cells("E1:E2")

        self.BOM_Sheet['A2'] = EATON_DESC
        self.BOM_Sheet['C3'] = "FILENAME"
        self.BOM_Sheet['D3'] = "REV"
        self.BOM_Sheet['E3'] = "DESCRIPTION"
        self.BOM_Sheet['E4'] = "BOM"
        self.BOM_Sheet['E5'] = "ASSEMBLY,SCHEMATIC"
        self.BOM_Sheet['E6'] = "PCB FILE"
        self.BOM_Sheet['E7'] = "PCB FAB"
        self.BOM_Sheet['E8'] = "PICK & PLACE FILE"
        self.BOM_Sheet['F1'] = "PICK & PLACE FILE"
        self.BOM_Sheet['E9'] = "HANDLING AND PROCESSING SPECS"
        self.BOM_Sheet['F1'] = "Description:"
        self.BOM_Sheet['E1'] = "Part Name: \nRevision: \nDesigner: \nEngineer: \nDate: \nCO #:"

        # Specify the start and end cells of the range
        start_cell = 'A1'
        end_cell = 'G9'

        # Wrap text in the specified range
        for row in self.BOM_Sheet[start_cell:end_cell]:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

        # Specify the picture name and desired dimensions
        picture_name = 'Picture 1'
        width_inches = 1.25
        height_inches = 0.29

    def AddHeader(self,Path):
        Template_Workbook = load_workbook(Path)
        Template_Sheet = Template_Workbook['Template']
        sheet = self.Workbook['BOM']

        for row in Template_Sheet.iter_rows(min_row=1, max_row=9, min_col=1, max_col=7):
            for cell in row:
                sheet[cell.coordinate].value = cell.value

    def MergeCells(self):
        # Choose the specific worksheet
        sheet = self.Workbook['BOM']  # Replace 'Sheet1' with your actual sheet name

        column_values = []

        StartRow = 7  # Replace with the specific row number you want to start from
        EndRow = sheet.max_row  # Replace with the actual end row if needed

        for i in range(StartRow, EndRow + 1):
            cell = sheet.cell(row=i, column=1)  # Replace 1 with the column number you want to extract from
            column_values.append(cell.value)

        sublists = {}

        for index, value in enumerate(column_values):
            if value in sublists:
                sublists[value].append(index)
            else:
                sublists[value] = [index]

        result = [sublists[key] for key in sublists]

        for column in range(1,6):
            for row in result:
                sheet.merge_cells(start_row=StartRow+row[0], start_column=column, end_row=StartRow+row[-1], end_column=column)

    def SaveDocument(self, Path):
        self.Workbook.save(Path)

if __name__ == '__main__':
    # Load the Excel file
    Path = '.\\test.xlsx'
    TemplatePath = '.\\EATON Template.xlsx'
    
    Excel = Template(Path)
    Excel.DeleteTabs()
    Excel.DeleteColumn()
    Excel.AddHeader1()
    Excel.Change_RowHeight()
    Excel.Change_ColumnWidth()
    Excel.mergecells()

    # Save the modified workbook
    Excel.SaveDocument('.\\test_modified.xlsx')
    print('Done')